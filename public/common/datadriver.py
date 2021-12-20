import os
import openpyxl
from public.common.readconfig import ReadConfig, filepath


def read_excel(xlsname, sheelname):
    """
    读取csv文件
    """
    read_config = ReadConfig(filepath)
    data_path = read_config.getValue('projectConfig', 'data_path')
    datapath = os.path.join(data_path, xlsname)
    csv = openpyxl.load_workbook(datapath)
    sheet = csv[sheelname]
    allList = []
    for row in range(2, sheet.max_row + 1):
        rowList = []
        for col in range(1, sheet.max_column + 1):
            rowList.append(sheet.cell(row, col).value)
        allList.append(rowList)
    return allList


""" 测试文件内容是否找到 """
if __name__ == '__main__':
    res = read_excel('user_data.xlsx', 'Sheet2')
    print(res)
